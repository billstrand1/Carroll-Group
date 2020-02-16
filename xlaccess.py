'''bonds.py

'''
from openpyxl import load_workbook
from pprint import pprint
import os

pprint ('openpyxl has been imported')

wb = load_workbook(filename = 'Handicap Index Course Handicap Report.xlsx')
pprint ('load_workbook worked into wb')

ws = wb['Sheet1']  #This may change with new File Export from Tap Forms
pprint ('sheet loaded into ws')

# Test cell objects vs. cell values
# print(ws['A2'].value)
# golfer_name = (ws['C2'].value)
# c = ws.cell(row=2, column=3) #cell object, identify by row/col numbers
# print(golfer_name, c.value)

player_dict = {}
#WORKS, print range of cells
for i in range(2, 21):
	golfer_name = (ws.cell(row=i, column=3).value)
	h_i = (ws.cell(row=i, column=4).value)
	player_dict[str(golfer_name)] = str(h_i)

pprint(player_dict)

#WORKS, print entire range based upon 'row i' not NULL
# i = 1
# while (ws.cell(row=i, column=1).value):
# 	print(i, ' ', ws.cell(row=i, column=1).value)
# 	i += 1

#WORKS, Total all Semiannual Income
# i = 2
# semi_annual = 0
# while (ws.cell(row=i, column=3).value):
# 	semi_annual += (ws.cell(row=i, column=3).value)
# 	i += 1
# print(semi_annual)

# #----------------START- Monthly Only-----------
# #WORKS, Total each monthly income from Jan - Jun
# monthly = {1 : 0, 2 : 0, 3 : 0, 4 : 0, 5 : 0, 6 : 0}

# i = 2
# while (ws.cell(row=i, column=5).value):
# 	month = ws.cell(row=i, column=5).value # 1 thru 6
# 	income = ws.cell(row=i, column=3).value # semi-annual income
# 	monthly[month] += income # add bonds's semi-annual
# 	i += 1

# #WORKS, append months 7 - 12 to monthly, same as 1-6
# for i in range (1, 7):
# 	monthly[i+6] = monthly[i]

# pprint(monthly)

# annual = 0
# for v in monthly.values():
# 	annual += v
# pprint (int(annual))
# #----------------FINISH- Monthly Only-----------


