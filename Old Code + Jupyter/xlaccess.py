# for testing of Handicap Index Excel Access

from openpyxl import load_workbook
from pprint import pprint
import os

pprint ('openpyxl has been imported')


wb = load_workbook(filename = 'Handicap Index Course Handicap Report.xlsx')
pprint ('load_workbook worked into wb')

ws = wb['Sheet1']  #This may change with new File Export from Tap Forms
pprint ('sheet loaded into ws')

# Test cell objects vs. cell values
# print(ws['A2'].value)
# golfer_name = (ws['C2'].value)
# c = ws.cell(row=2, column=3) #cell object, identify by row/col numbers
# print(golfer_name, c.value)

player_dict = {}
#WORKS, print range of cells
for i in range(2, 21):
	golfer_name = (ws.cell(row=i, column=3).value)
	h_i = (ws.cell(row=i, column=4).value)
	player_dict[str(golfer_name)] = str(h_i)

pprint(player_dict)


