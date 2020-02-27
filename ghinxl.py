# Updated for Excel utilization

from tabulate import tabulate
from datetime import datetime
from openpyxl import load_workbook
from pprint import pprint
import os

# contains the name dictionary and Player class, course handicap information, player status
from players import *

# cd documents/github/Carroll-Group

#--------------------Functions -Excel HI's-----------------------------

#When using the Excel File from Golf Genius
def get_indexes_from_xl_file():
	wb = load_workbook(filename = 'Handicap Index Course Handicap Report.xlsx')
	# print(wb)
	ws = wb['Sheet1']
	# print(ws)

	#Retrieve the player's indexes, compute handicaps
	for i in range(2, 21):
		print(i)
		golfer_name = (ws.cell(row=i, column=3).value)
		print(golfer_name)
		# h_i = (float(ws.cell(row=i, column=4).value))


		#Update handicap info in Player Calss
		for player in player_list:
			if golfer_name == player.ghin_name:
				player.h_i = h_i
				player.class_tpc_white_70()
				player.class_tpc_white_72()
				player.class_cwv_white_71()
				# print (f" {golfer_name}, hi: {player.h_i}, tpc 70: {player.handicap_tpc_70}, tpc 72: {player.handicap_tpc_72} cwv: {player.handicap_cwv_71}")

#------------------PRINT OUT TABLE OF INDEXES / HANDICAPS / STROKES OFF LOW HANDICAP----------
def print_results(tpc_min_70,tpc_min_72, cwv_min):  
	today = datetime.now()
	today = today.strftime("%B %d, %Y")

	results_list = []

	choice = input ("[T]PC or [C]WV > ")

	if choice == 'T' or choice == 't':
		for player in player_list:
			if player.playing:
				results = [player.signup_name, player.h_i, player.handicap_tpc_70, player.handicap_tpc_70 - tpc_min_70, player.handicap_tpc_72, player.handicap_tpc_72 - tpc_min_72]
				results_list.append(results)

		results_list.sort()
		results_list.insert(0,["   ", "   ",  "TPC", "TPC-70", "TPC", "TPC-72"])
		results_list.insert(1,["Name", "Index","(70)", "Strks", "(72)", "Strks"])
		column_print_alignment = list(("right","right","right","right", "right", "right"))
		course_info = f"TPC 70: CR = {tpc_rating_white_70} | SR = {tpc_slope_white_70} | Par = {tpc_hcp_70} \n"
		course_info = course_info + f"TPC 72: CR = {tpc_rating_white_72} | SR = {tpc_slope_white_72} | Par = {tpc_hcp_72}"

	elif choice == 'C' or choice == 'c':
		for player in player_list:
			if player.playing:
				results = [player.signup_name, player.h_i, player.handicap_cwv_71, player.handicap_cwv_71-cwv_min]
				results_list.append(results)

		results_list.sort()
		results_list.insert(0,["   ", "   ", "CWV HCP", "CWV"])
		results_list.insert(1,["Name", "Index","Par 71", "Strokes"])
		column_print_alignment = list(("right","right","right","right"))
		course_info = f"CWV: CR = {cwv_rating_white_71} | SR = {cwv_slope_white_71} | Par = {cwv_hcp_71}"

	else: print('Bad Choice')
	
	#Finish printout:
	print('-------------------------------------------------------')
	print("\n \n Today's date:", today)
	print (tabulate(results_list, tablefmt='fancy_grid', colalign=column_print_alignment))
	print(course_info)
	print ('Course Handicap = (H.I. x SR / 113) + (CR - Par)')
	print('--------------------------------------------------------')
	for player in player_list:
		if player.playing:
			print (player.email)
	print("\n\n")
	#END-----------------PRINT OUT TABLE OF INDEXES / HANDICAPS / STROKES OFF LOW HANDICAP----------		

def main():
	#Update the indexes of all players from the excel file
	get_indexes_from_xl_file()

	#accumulate list of handicaps for sorting purposes
	tpc_list_70 = []
	tpc_list_72 = [] 
	cwv_list = [] 

	#TODO:  Complete the Smartsheet sign-up list to update player.playing

	for player in player_list:
		if player.playing:
			tpc_list_70.append(player.handicap_tpc_70)
			tpc_list_72.append(player.handicap_tpc_72)
			cwv_list.append(player.handicap_cwv_71)

	#determine lowest handicaps
	tpc_list_70.sort()
	tpc_min_70 = tpc_list_70[0] #lowest TPC70 handicap
	print(f"TPC 70 Min: {tpc_min_70}")

	tpc_list_72.sort()
	tpc_min_72 = tpc_list_72[0] #lowest TPC72 handicap
	print(f"TPC 72 Min: {tpc_min_72}")

	cwv_list.sort()
	cwv_min = cwv_list[0] #lowest CWV handicap
	print(f"CWV Min: {cwv_min}")

	print_results(tpc_min_70,tpc_min_72, cwv_min)

#--------------------START CODE EXECUTION-----------------------------

# <section class="golfer_lookup_section"> is copied to text file...

main()
